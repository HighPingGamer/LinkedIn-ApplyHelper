"""Collects one row per job the agent acted on and writes a formatted Excel sheet
at the end of the session."""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADERS = ["Timestamp", "Company / Role", "Location", "Salary (AED)",
           "Fit Score", "Status", "Resume Modified", "Application Link", "Note"]


class SessionLog:
    def __init__(self):
        self.rows = []

    def add(self, company_role, location, salary, fit, status, resume_modified, link, note):
        self.rows.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            company_role, location,
            salary if salary is not None else "",
            fit, status,
            "Yes" if resume_modified else "No",
            link, note,
        ])

    def write(self, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"

        head_fill = PatternFill("solid", fgColor="1F3864")
        head_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(HEADERS)
        for i, _ in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=i)
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            c.border = border

        status_colors = {"submitted": "C6EFCE", "skipped": "FFEB9C", "failed": "FFC7CE"}
        for r in self.rows:
            ws.append(r)
            row_idx = ws.max_row
            status = str(r[5]).lower()
            fill = status_colors.get(status)
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col == 6 and fill:
                    cell.fill = PatternFill("solid", fgColor=fill)

        widths = [16, 34, 18, 12, 9, 12, 14, 40, 44]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"

        # Summary line at the bottom
        submitted = sum(1 for r in self.rows if str(r[5]).lower() == "submitted")
        skipped = sum(1 for r in self.rows if str(r[5]).lower() == "skipped")
        failed = sum(1 for r in self.rows if str(r[5]).lower() == "failed")
        ws.append([])
        ws.append(["", f"Submitted: {submitted}", "", "", "",
                   f"Skipped: {skipped}", f"Failed: {failed}"])
        output_path = Path(path)
        try:
            wb.save(output_path)
        except PermissionError:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback = output_path.with_name(f"{output_path.stem}_{stamp}{output_path.suffix}")
            wb.save(fallback)
            output_path = fallback
        return submitted, skipped, failed, str(output_path)
